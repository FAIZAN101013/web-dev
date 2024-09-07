import  openpyxl as xl
from openpyxl.chart import BarChart,Reference




wb = xl.load_workbook('transactions.xlsx')
with open(wb ,"w") as file:
sheet = wb['Sheet1'] 
cell = sheet['a1']
cell = sheet.cell(1,1)



for row in range(2,sheet.max_row + 1):
   cell = sheet.cell(row ,3)
   c_price = cell.value * 0.9
   
   c_p_cell = sheet.cell(row ,4)
   c_p_cell.value = c_price
   
v = Reference(sheet,
                min_row = 2,
                max_row = sheet.max_row,
                max_col= 4,
                min_col=4)
chart = BarChart()
chart.add_data(v)
sheet.add_chart(chart,'e2')


wb.save('t2.xlsx')
    